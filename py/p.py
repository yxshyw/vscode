import openpyxl
path = r'C:\\Users\\yxshyw\\Desktop\\t.xlsx'

wb = openpyxl.load_workbook(path)
ws = wb.get_sheet_by_name('geg')

def read_merge_cell():
    merge_all_list=[]
    for m_area in ws.merged_cells:
        # 合并单元格的起始行坐标、终止行坐标
        r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
        # print(r1, r2, c1, c2)
        if (r1 != r2 and c1 != c2):
            row_col = [(x, y) for x in range(r1, r2 + 1) for y in range(c1, c2 + 1)]
            merge_all_list.append(row_col)
            # print(r1, r2, c1, c2)
        elif (r1 == r2 and c1 != c2):  # or (r1 != r2 and c1 == c2):
            col = [(r1, n) for n in range(c1, c2 + 1)]
            # print(col)
            merge_all_list.append(col)
        elif (r1 != r2 and c1 == c2):
            row = [(m, c1) for m in range(r1, r2 + 1)]
            merge_all_list.append(row)
    return merge_all_list

def read_excel():
    row = []
    for i in range(1,ws.max_row+1):
        tmp = []
        for j in range(1,ws.max_column+1):
            for k in read_merge_cell():
                # print(i,j)
                if (i, j) in k:
                    # print(ws.cell(*k[0]).value)
                    tmp.append(ws.cell(*k[0]).value)
                else:
                    # print(ws.cell(i, j).value)
                    tmp.append(ws.cell(i, j).value)
                break
        row.append(tmp)
    return row

print(read_excel())